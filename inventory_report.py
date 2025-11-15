from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import List, Optional
import re

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SOURCE_DIR = Path(
    r"C:\Users\Alexandru\Documents\DOCUMENTE CRAMA\pictograme brand\facturi furnizori"
)
OUTPUT_PATH = Path("Inventory Report.xlsx")

# Known units of measurement
UNIT_CANDIDATES = {
    "BUC", "BUCS", "BUC.", "KG", "KGM", "KGS", "KGF", "KGD",
    "L", "LIT", "LTS", "ML", "GR", "G", "MC", "PCE", "PCS",
    "SET", "KIT", "BAX", "ST", "ROL", "M", "MM", "CM", "MET",
    "H87", "TA8", "TA7", "TA6", "H89",  # Common packaging codes
}


@dataclass
class ProductEntry:
    product: str
    unit: str
    quantity: float
    source_file: str
    doc_type: str
    doc_date: Optional[date]
    doc_number: str
    price: Optional[float]
    notes: str
    color: Optional[str] = None
    capacity_ml: Optional[float] = None
    weight_g: Optional[float] = None
    height_mm: Optional[float] = None
    diameter_mm: Optional[float] = None
    closure_mm: Optional[float] = None
    pallet_dimensions: Optional[str] = None


def normalize_number(token: str) -> float:
    """Convert various number formats to float."""
    cleaned = (
        token.replace("\u202f", "")
        .replace(" ", "")
        .replace("\xa0", "")
        .replace(",", ".")
    )
    return float(cleaned)


def parse_document_date(text: str, path: Path) -> Optional[date]:
    """Extract document date from text or filename."""
    patterns = ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y")
    matches = re.findall(r"\b(\d{2}[./-]\d{2}[./-]\d{4})\b", text)
    for match in matches:
        for pattern in patterns:
            try:
                return datetime.strptime(match, pattern).date()
            except ValueError:
                continue
    # Try to parse from filename
    file_match = re.search(r"__(\d{2})-(\d{2})-(\d{4})__", path.name)
    if file_match:
        day, month, year = file_match.groups()
        return date(int(year), int(month), int(day))
    return datetime.fromtimestamp(path.stat().st_mtime).date()


def parse_document_number(text: str, path: Path) -> str:
    """Extract the document number."""
    match = re.search(r"Nr\.?\s+doc\.?\s+([A-Z0-9/-]+)", text, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    file_match = re.search(r"__([A-Z0-9-]+)_index", path.name, flags=re.IGNORECASE)
    if file_match:
        return file_match.group(1)
    return path.stem


def infer_document_type(path: Path) -> str:
    """Determine if it's a delivery note or invoice."""
    lower_name = path.name.lower()
    if "nota_intrare" in lower_name or "nota intrare" in lower_name:
        return "Delivery Note"
    if "factura" in lower_name:
        return "Invoice"
    return "Document"


def extract_product_line(line: str) -> Optional[tuple]:
    """
    Parse a product line from invoice/delivery note.
    Expected format: # ProductName U.M. Qty Price ...
    Returns: (product_name, unit, quantity, price) or None
    """
    tokens = line.split()
    if len(tokens) < 4:
        return None
    
    # First token must be a number (line number)
    if not tokens[0].rstrip(".").isdigit():
        return None
    
    # Find the unit: look for known unit codes
    unit_idx = None
    unit = ""
    for idx, token in enumerate(tokens[1:], start=1):
        normalized = token.strip("().,").upper()
        if normalized in UNIT_CANDIDATES:
            unit_idx = idx
            unit = normalized
            break
    
    if unit_idx is None or unit_idx <= 1:
        return None
    
    # Everything between line number and unit is the product name
    product_tokens = tokens[1:unit_idx]
    product_name = " ".join(product_tokens).strip()
    
    if not product_name or not any(c.isalpha() for c in product_name):
        return None
    
    # Quantity is immediately after unit
    if unit_idx + 1 >= len(tokens):
        return None
    
    qty_token = tokens[unit_idx + 1]
    try:
        quantity = normalize_number(qty_token)
    except ValueError:
        return None
    
    # Price is after quantity (first numeric value after qty)
    price = None
    if unit_idx + 2 < len(tokens):
        for token in tokens[unit_idx + 2:]:
            try:
                price = normalize_number(token)
                break
            except ValueError:
                continue
    
    return (product_name, unit, quantity, price)


NOTE_KEYWORDS = (
    "lot",
    "expir",
    "observ",
    "serie",
    "batch",
    "culoare",
    "capacitate",
    "greutate",
    "inaltime",
    "diametru",
    "inchidere",
    "paleti",
    "dimensiuni",
)


def parse_attribute_line(line: str, attrs: dict[str, Optional[str | float]]) -> bool:
    """Parse attribute details from a note line."""
    updated = False

    def parse_number(value: str) -> Optional[float]:
        try:
            return normalize_number(value)
        except ValueError:
            return None

    color_match = re.search(r"Culoare\s*:\s*([^;]+)", line, flags=re.IGNORECASE)
    if color_match:
        attrs["color"] = color_match.group(1).strip()
        updated = True

    capacity_match = re.search(
        r"Capacitate\s*(?:\[[^\]]*\])?\s*:\s*([0-9.,]+)", line, flags=re.IGNORECASE
    )
    if capacity_match:
        attrs["capacity_ml"] = parse_number(capacity_match.group(1))
        updated = True

    weight_match = re.search(
        r"Greutate\s*(?:\[[^\]]*\])?\s*:\s*([0-9.,]+)", line, flags=re.IGNORECASE
    )
    if weight_match:
        attrs["weight_g"] = parse_number(weight_match.group(1))
        updated = True

    height_match = re.search(
        r"Înălțime|Inaltime\s*(?:\[[^\]]*\])?\s*:\s*([0-9.,]+)",
        line,
        flags=re.IGNORECASE,
    )
    if height_match:
        attrs["height_mm"] = parse_number(height_match.group(1))
        updated = True

    diameter_match = re.search(
        r"Diametru\s*(?:\[[^\]]*\])?\s*:\s*([0-9.,]+)", line, flags=re.IGNORECASE
    )
    if diameter_match:
        attrs["diameter_mm"] = parse_number(diameter_match.group(1))
        updated = True

    closure_match = re.search(
        r"Inchidere\s*(?:\[[^\]]*\])?\s*:\s*([0-9.,]+)", line, flags=re.IGNORECASE
    )
    if closure_match:
        attrs["closure_mm"] = parse_number(closure_match.group(1))
        updated = True

    pallet_match = re.search(
        r"Dimensiuni\s*:\s*([^;]+)", line, flags=re.IGNORECASE
    )
    if pallet_match:
        attrs["pallet_dimensions"] = pallet_match.group(1).strip()
        updated = True

    return updated


def extract_entries_from_file(path: Path) -> List[ProductEntry]:
    """Extract all product entries from a PDF file."""
    entries: List[ProductEntry] = []

    with pdfplumber.open(path) as pdf:
        page_texts = [page.extract_text() or "" for page in pdf.pages]

    combined_text = "\n".join(page_texts)
    doc_date = parse_document_date(combined_text, path)
    doc_type = infer_document_type(path)
    doc_number = parse_document_number(combined_text, path)

    for text in page_texts:
        if not text:
            continue

        lines = text.splitlines()
        idx = 0
        while idx < len(lines):
            line = lines[idx].strip()
            if not line:
                idx += 1
                continue

            result = extract_product_line(line)
            if not result:
                idx += 1
                continue

            product_name, unit, quantity, price = result

            notes: List[str] = []
            attributes: dict[str, Optional[str | float]] = {
                "color": None,
                "capacity_ml": None,
                "weight_g": None,
                "height_mm": None,
                "diameter_mm": None,
                "closure_mm": None,
                "pallet_dimensions": None,
            }
            lookahead = idx + 1
            while lookahead < len(lines):
                peek = lines[lookahead].strip()
                if not peek:
                    lookahead += 1
                    continue
                if peek[0].isdigit():
                    break
                lowered = peek.lower()
                if lowered.startswith(NOTE_KEYWORDS):
                    if not parse_attribute_line(peek, attributes):
                        notes.append(peek)
                    lookahead += 1
                    continue
                break

            entries.append(
                ProductEntry(
                    product=product_name,
                    unit=unit,
                    quantity=quantity,
                    source_file=path.name,
                    doc_type=doc_type,
                    doc_date=doc_date,
                    doc_number=doc_number,
                    price=price,
                    notes="; ".join(notes),
                    color=attributes["color"],
                    capacity_ml=attributes["capacity_ml"],
                    weight_g=attributes["weight_g"],
                    height_mm=attributes["height_mm"],
                    diameter_mm=attributes["diameter_mm"],
                    closure_mm=attributes["closure_mm"],
                    pallet_dimensions=attributes["pallet_dimensions"],
                )
            )
            idx = lookahead

    return entries


def aggregate_entries(entries: List[ProductEntry]) -> List[dict]:
    """
    Consolidate entries by product and unit.
    For each product, keep the last (by date) price, doc type, date, etc.
    """
    aggregate: dict[tuple[str, str], dict] = {}
    
    def sort_key(entry: ProductEntry):
        date_val = entry.doc_date or date(1900, 1, 1)
        return (date_val, entry.source_file)
    
    for entry in entries:
        key = (entry.product.lower(), entry.unit.upper())
        
        if key not in aggregate:
            aggregate[key] = {
                "product": entry.product,
                "unit": entry.unit,
                "quantity": 0.0,
                "last_entry": None,
            }
        
        aggregate[key]["quantity"] += entry.quantity
        
        last_entry = aggregate[key]["last_entry"]
        if last_entry is None or sort_key(entry) >= sort_key(last_entry):
            aggregate[key]["last_entry"] = entry
    
    # Build final rows
    aggregated_rows: List[dict] = []
    for record in aggregate.values():
        last_entry = record["last_entry"]
        aggregated_rows.append(
            {
                "product": record["product"],
                "unit": record["unit"],
                "quantity": record["quantity"],
                "last_price": last_entry.price if last_entry else None,
                "document_type": last_entry.doc_type if last_entry else "",
                "document_date": last_entry.doc_date if last_entry else None,
                "document_number": last_entry.doc_number if last_entry else "",
                "source_file": last_entry.source_file if last_entry else "",
                "notes": last_entry.notes if last_entry else "",
                "color": last_entry.color if last_entry else None,
                "capacity_ml": last_entry.capacity_ml if last_entry else None,
                "weight_g": last_entry.weight_g if last_entry else None,
                "height_mm": last_entry.height_mm if last_entry else None,
                "diameter_mm": last_entry.diameter_mm if last_entry else None,
                "closure_mm": last_entry.closure_mm if last_entry else None,
                "pallet_dimensions": last_entry.pallet_dimensions if last_entry else None,
            }
        )
    
    return sorted(aggregated_rows, key=lambda x: x["product"].lower())


def format_sheet(ws, report_date: date):
    """Format the header and columns."""
    ws.title = "Inventory"
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Inventory Report as of {report_date.strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    headers = [
        "No.",
        "Product",
        "Unit",
        "Total Qty",
        "Stock (as of report)",
        "Last Price (RON)",
        "Document Type",
        "Document Date",
        "Document No.",
        "Color",
        "Capacity (ml)",
        "Weight (g)",
        "Height (mm)",
        "Diameter (mm)",
        "Closure (mm)",
        "Pallet Dimensions",
        "Product Notes",
        "Physical Inventory (to fill in)",
        "Remarks",
    ]
    ws.append(headers)
    
    header_row = 2
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    column_widths = [
        6,
        40,
        10,
        12,
        15,
        14,
        16,
        14,
        14,
        16,
        14,
        14,
        14,
        14,
        18,
        30,
        25,
        20,
    ]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(ord("A") + idx - 1)].width = width


def write_excel_report(aggregated_rows: List[dict], all_entries: List[ProductEntry], output_path: Path) -> Path:
    """Write data to Excel workbook."""
    wb = Workbook()
    ws = wb.active
    today = date.today()
    format_sheet(ws, today)
    
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    
    start_row = 3
    column_count = len(ws[2])
    
    for idx, row in enumerate(aggregated_rows, start=1):
        excel_row = start_row + idx - 1
        ws.cell(row=excel_row, column=1, value=idx)
        ws.cell(row=excel_row, column=2, value=row["product"])
        ws.cell(row=excel_row, column=3, value=row["unit"])
        ws.cell(row=excel_row, column=4, value=row["quantity"])
        ws.cell(row=excel_row, column=5, value=row["quantity"])
        
        price_cell = ws.cell(row=excel_row, column=6, value=row.get("last_price"))
        if price_cell.value is not None:
            price_cell.number_format = "#,##0.00"
        
        ws.cell(row=excel_row, column=7, value=row.get("document_type", ""))
        
        date_cell = ws.cell(row=excel_row, column=8, value=row.get("document_date"))
        if date_cell.value:
            date_cell.number_format = "DD.MM.YYYY"
        
        ws.cell(row=excel_row, column=9, value=row.get("document_number", ""))
        ws.cell(row=excel_row, column=10, value=row.get("color", ""))
        ws.cell(row=excel_row, column=11, value=row.get("capacity_ml"))
        ws.cell(row=excel_row, column=12, value=row.get("weight_g"))
        ws.cell(row=excel_row, column=13, value=row.get("height_mm"))
        ws.cell(row=excel_row, column=14, value=row.get("diameter_mm"))
        ws.cell(row=excel_row, column=15, value=row.get("closure_mm"))
        ws.cell(row=excel_row, column=16, value=row.get("pallet_dimensions", ""))
        ws.cell(row=excel_row, column=17, value=row.get("notes", ""))
        ws.cell(row=excel_row, column=18, value="")
        ws.cell(row=excel_row, column=19, value="")

        for numeric_col in (11, 12, 13, 14, 15):
            num_cell = ws.cell(row=excel_row, column=numeric_col)
            if num_cell.value is not None:
                num_cell.number_format = "#,##0.00"
        
        for col in range(1, column_count + 1):
            cell = ws.cell(row=excel_row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        ws.row_dimensions[excel_row].height = 22
    
    # Total row
    total_row = start_row + len(aggregated_rows)
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=4, value=f"=SUM(D{start_row}:D{total_row-1})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E{start_row}:E{total_row-1})")
    
    for col in range(1, column_count + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="double"), left=thin, right=thin, bottom=thin)
    
    # Notes row
    notes_row = total_row + 2
    ws.merge_cells(f"A{notes_row}:K{notes_row}")
    ws.cell(
        row=notes_row,
        column=1,
        value=f"Generated from {len(all_entries)} items extracted from invoices in: {SOURCE_DIR}",
    )
    ws.cell(row=notes_row, column=1).alignment = Alignment(wrap_text=True)
    
    # Try to save, fallback if locked
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        wb.save(fallback)
        print(f"WARNING: {output_path} is open. Saved to: {fallback}")
        return fallback


def main():
    if not SOURCE_DIR.exists():
        raise SystemExit(f"Directory not found: {SOURCE_DIR}")
    
    # Process only invoices (e-Factura files), not delivery notes (Nota intrare)
    all_pdf_files = sorted(SOURCE_DIR.glob("*.pdf"))
    pdf_files = [f for f in all_pdf_files if "factura" in f.name.lower()]
    
    if not pdf_files:
        raise SystemExit(f"No invoice files found in {SOURCE_DIR}")
    
    all_entries: List[ProductEntry] = []
    for pdf_path in pdf_files:
        entries = extract_entries_from_file(pdf_path)
        all_entries.extend(entries)
        print(f"{pdf_path.name}: {len(entries)} items extracted")
    
    if not all_entries:
        raise SystemExit("No products found in PDF files.")
    
    aggregated = aggregate_entries(all_entries)
    print(f"\nTotal unique products: {len(aggregated)}")
    print(f"Total line items processed: {len(all_entries)}")
    
    saved_path = write_excel_report(aggregated, all_entries, OUTPUT_PATH)
    print(f"\nReport saved: {saved_path}")


if __name__ == "__main__":
    main()
